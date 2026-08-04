#!/usr/bin/env python3
"""
Route Definitions: Call Routes
Exposes REST endpoints for triggering calls, fetching status, and receiving callbacks.
Generates comprehensive OpenAPI Swagger schemas.
"""

from fastapi import APIRouter, HTTPException, Depends, status
from pydantic import BaseModel, Field
from typing import Dict, Any, Optional, Union, List
from controllers import call_controller

router = APIRouter(
    prefix="/api/v1/calls",
    tags=["Call Management"],
    responses={
        401: {"description": "Unauthorized Access - Bearer token missing or invalid"},
        500: {"description": "Internal Server Error"}
    }
)


# === Pydantic Input Schemas for Request Validation ===

class OutboundCallRequest(BaseModel):
    phone_number: str = Field(
        ..., 
        example="+919876543210", 
        description="Target phone number string or comma-separated list of phone numbers (e.g. '+919876543210, +919876543211')."
    )
    customer_name: str = Field(
        "Customer", 
        example="John Doe", 
        description="Name of the customer being called to personalize synthesized greeting."
    )
    enterprise_id: Optional[str] = Field(
        "ent_default",
        example="ent_admin_101",
        description="Enterprise ID to verify and associate the call on behalf of an enterprise admin."
    )
    agent_id: Optional[str] = Field(
        "default",
        example="agent_sales_01",
        description="Agent ID associated with this call for specific AI voice configuration."
    )
    campaign_id: Optional[str] = Field(
        None,
        example="cmp_20260731_a1b2c3d4",
        description="Campaign ID. Automatically generated if omitted."
    )
    context: Optional[Dict[str, Any]] = Field(
        None,
        example={"source": "HubSpot"},
        description="Arbitrary dictionary context to trace conversation history and CRM updates."
    )

class WebhookCallbackPayload(BaseModel):
    CallSid: str = Field(..., example="ex_call_8e90810557fc4dc4ab5c04", description="Exotel Call identifier.")
    EventType: str = Field(..., example="call.completed", description="The nature of the callback event.")
    Duration: Optional[int] = Field(None, example=45, description="Call duration in seconds.")
    CustomData: Optional[str] = Field(None, description="Optional raw serialised context passed during trigger.")

# === Pydantic Output Schemas for Swagger Documentation ===

class CallActionResponse(BaseModel):
    success: bool = Field(..., json_schema_extra={"example": True})
    call_sid: Optional[str] = Field(None, json_schema_extra={"example": "ex_call_8e90810557fc4dc4ab5c04"})
    call_sids: Optional[List[str]] = Field(None, json_schema_extra={"example": ["ex_call_8e90810557fc4dc4ab5c04"]})
    enterprise_id: Optional[str] = Field(None, json_schema_extra={"example": "ent_admin_101"})
    agent_id: Optional[str] = Field(None, json_schema_extra={"example": "agent_sales_01"})
    campaign_id: Optional[str] = Field(None, json_schema_extra={"example": "cmp_20260731_a1b2c3d4"})
    status: Optional[str] = Field(None, json_schema_extra={"example": "initiated"})
    message: Optional[str] = Field(None, json_schema_extra={"example": "Outbound call request initiated successfully."})
    error: Optional[str] = Field(None, json_schema_extra={"example": None})

class CallStatusDetailsResponse(BaseModel):
    success: bool = Field(..., json_schema_extra={"example": True})
    call_sid: str = Field(..., json_schema_extra={"example": "ex_call_8e90810557fc4dc4ab5c04"})
    status: str = Field(..., json_schema_extra={"example": "completed"})
    duration: Optional[int] = Field(45, json_schema_extra={"example": 45})
    direction: str = Field("outbound", json_schema_extra={"example": "outbound"})
    from_number: Optional[str] = Field(None, json_schema_extra={"example": "+918047190000"})
    to_number: Optional[str] = Field(None, json_schema_extra={"example": "+919876543210"})
    start_time: Optional[str] = Field(None, json_schema_extra={"example": "2026-05-22 10:17:41"})
    price: Optional[Union[str, float]] = Field(None, json_schema_extra={"example": "0.50"})
    error: Optional[str] = Field(None, json_schema_extra={"example": None})

class CampaignCallDetail(BaseModel):
    call_sid: Optional[str] = Field(None, json_schema_extra={"example": "ex_call_8e90810557fc4dc4ab5c04"})
    phone_number: Optional[str] = Field(None, json_schema_extra={"example": "+919876543210"})
    customer_name: Optional[str] = Field(None, json_schema_extra={"example": "John Doe"})
    status: Optional[str] = Field(None, json_schema_extra={"example": "completed"})
    duration: Optional[int] = Field(None, json_schema_extra={"example": 45})
    timestamp: Optional[float] = Field(None, json_schema_extra={"example": 1785228102.8})

class CampaignSummary(BaseModel):
    campaign_id: str = Field(..., json_schema_extra={"example": "cmp_20260731_a1b2c3d4"})
    enterprise_id: str = Field(..., json_schema_extra={"example": "ent_admin_101"})
    agent_id: str = Field(..., json_schema_extra={"example": "agent_sales_01"})
    total_calls: int = Field(..., json_schema_extra={"example": 10})
    completed_calls: int = Field(..., json_schema_extra={"example": 8})
    failed_calls: int = Field(..., json_schema_extra={"example": 1})
    initiated_calls: int = Field(..., json_schema_extra={"example": 1})
    in_progress_calls: int = Field(..., json_schema_extra={"example": 0})
    first_call_timestamp: Optional[float] = Field(None, json_schema_extra={"example": 1785228102.8})
    latest_call_timestamp: Optional[float] = Field(None, json_schema_extra={"example": 1785229000.0})
    calls: List[CampaignCallDetail] = Field(default_factory=list)

class CampaignsResponse(BaseModel):
    success: bool = Field(..., json_schema_extra={"example": True})
    enterprise_id: str = Field(..., json_schema_extra={"example": "ent_admin_101"})
    agent_id: Optional[str] = Field(None, json_schema_extra={"example": "agent_sales_01"})
    total_campaigns: int = Field(..., json_schema_extra={"example": 1})
    campaigns: List[CampaignSummary] = Field(default_factory=list)

# === API Endpoint Route Mappings ===

@router.post(
    "/outbound",
    response_model=CallActionResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger Outbound Lead Call",
    description="Dial an outbound phone call (or batch of numbers) via the Exotel gateway on behalf of an Enterprise Admin & Agent. Auto-generates campaign ID if omitted."
)
async def trigger_call(payload: OutboundCallRequest):
    """Trigger an outbound call using Exotel gateway REST API."""
    target_phones = payload.phone_number
    if not target_phones or not str(target_phones).strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Target phone number must be provided in 'phone_number'."
        )

    result = await call_controller.initiate_outbound_call(
        phone_number=target_phones,
        customer_name=payload.customer_name,
        enterprise_id=payload.enterprise_id,
        agent_id=payload.agent_id,
        campaign_id=payload.campaign_id,
        context=payload.context
    )
    
    if not result.get("success"):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=result.get("error", "An error occurred initiating outbound call")
        )
        
    return result

@router.get(
    "/status/{call_sid}",
    response_model=CallStatusDetailsResponse,
    summary="Retrieve Outbound Call Status",
    description="Inspect the real-time state, connection durations, billing costs, and outcomes of a call using the Exotel session SID."
)
async def get_status(call_sid: str):
    """Retrieve full Call details from telephony service."""
    result = await call_controller.fetch_call_status(call_sid)
    
    if not result.get("success"):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=result.get("error", "Call SID not found or unavailable")
        )
        
    return result

from fastapi import Request

@router.post(
    "/webhook",
    summary="Telephony Callback Handler",
    description="Receive callbacks from Exotel gateways to audit call lifecycle (ringing, answers, timeouts, disconnects)."
)
async def call_webhook(request: Request):
    """Register incoming webhooks from carrier callback nodes."""
    content_type = request.headers.get("content-type", "")
    payload = {}
    if "application/x-www-form-urlencoded" in content_type:
        form_data = await request.form()
        payload = dict(form_data)
    else:
        try:
            payload = await request.json()
        except Exception:
            form_data = await request.form()
            payload = dict(form_data)
            
    logger.info(f"📥 Received Exotel webhook payload: {payload}")
    result = await call_controller.process_telephony_webhook(payload)
    
    if not result.get("success"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result.get("error")
        )
        
    return result

# === New Endpoints for Dashboard Outbound Call Status & Batch Calling ===

import logging
import csv
import io
from fastapi import UploadFile, File

logger = logging.getLogger(__name__)

@router.get(
    "/logs",
    summary="Get Filtered Call Logs",
    description="Retrieve historical call logs filtered by enterprise_id, agent_id, phone_number, and direction."
)
async def get_call_logs(
    enterprise_id: Optional[str] = Query(None, description="Enterprise ID to filter call logs"),
    agent_id: Optional[str] = Query(None, description="Agent ID (or Mongo _id) to filter call logs"),
    phone_number: Optional[str] = Query(None, description="Customer or Virtual Phone Number to filter logs"),
    direction: Optional[str] = Query(None, description="Call direction filter ('inbound' or 'outbound')"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(50, ge=1, le=500, description="Records per page"),
    x_enterprise_id: Optional[str] = Header(None, alias="x-enterprise-id")
):
    ent_id = (enterprise_id or x_enterprise_id or "").strip()
    try:
        from core.mongo_manager import mongo_db
        from bson import ObjectId
        import re
        
        if mongo_db.client is None:
            return {"success": True, "total": 0, "page": page, "limit": limit, "logs": []}
            
        db = mongo_db.client.get_default_database()
        calls_coll = db['Agent_Stream_CallsLogs']
        
        query_clauses = []
        
        if ent_id:
            query_clauses.append({
                "$or": [
                    {"enterprise_id": ent_id},
                    {"enterprise": ent_id},
                    {"company_id": ent_id},
                    {"createdBy": ent_id}
                ]
            })
            
        if agent_id and agent_id.strip():
            target_agent = agent_id.strip()
            agent_or = [
                {"agentId": target_agent},
                {"agent_id": target_agent}
            ]
            if ObjectId.is_valid(target_agent):
                agent_or.append({"_id": ObjectId(target_agent)})
            query_clauses.append({"$or": agent_or})
            
        if phone_number and phone_number.strip():
            clean_digits = re.sub(r'\D', '', str(phone_number))[-10:]
            if clean_digits:
                rgx = {"$regex": clean_digits}
                query_clauses.append({
                    "$or": [
                        {"phone_number": rgx},
                        {"phoneNumber": rgx},
                        {"to_number": rgx},
                        {"from_number": rgx}
                    ]
                })
                
        if direction and direction.strip().lower() in ["inbound", "outbound"]:
            query_clauses.append({"direction": direction.strip().lower()})
            
        final_filter = {}
        if len(query_clauses) == 1:
            final_filter = query_clauses[0]
        elif len(query_clauses) > 1:
            final_filter = {"$and": query_clauses}
            
        skip = (page - 1) * limit
        total_count = await calls_coll.count_documents(final_filter)
        cursor = calls_coll.find(final_filter).sort("timestamp", -1).skip(skip).limit(limit)
        
        logs = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            logs.append(doc)
            
        return {
            "success": True,
            "total": total_count,
            "page": page,
            "limit": limit,
            "logs": logs
        }
    except Exception as e:
        logger.error(f"Error querying call logs: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Internal Server Error: {str(e)}"
        )

@router.get(
    "/outbound",
    summary="Get Outbound Call List",
    description="Retrieves a list of all outbound calls from MongoDB to display their status (initiated, ringing, completed, failed) in the admin panel."
)
async def get_outbound_calls():
    try:
        from core.mongo_manager import mongo_db
        if mongo_db.client is None:
            return {"success": True, "calls": []}
        db = mongo_db.client.get_default_database()
        cursor = db['outbound_calls'].find().sort("timestamp", -1)
        calls = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            calls.append(doc)
        return {"success": True, "calls": calls}
    except Exception as e:
        logger.error(f"Error fetching outbound calls: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Internal Server Error: {str(e)}"
        )

from fastapi import Query, UploadFile, File
import uuid
from datetime import datetime

@router.post(
    "/outbound/bulk",
    summary="Bulk Outbound Calls via CSV or Excel",
    description="Upload a CSV or Excel (.xlsx) file containing phone_number and customer_name to initiate a batch of outbound calls."
)
async def trigger_bulk_calls(
    file: UploadFile = File(...),
    enterprise_id: Optional[str] = Query(None, description="Enterprise ID for verification on behalf of enterprise admin"),
    agent_id: Optional[str] = Query(None, description="Agent ID for AI configuration"),
    campaign_id: Optional[str] = Query(None, description="Campaign ID (Auto-generated if omitted)")
):
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported.")
    
    # Auto-generate a single campaign_id for this bulk batch if omitted
    batch_campaign_id = campaign_id or f"cmp_bulk_{datetime.now().strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"

    try:
        contents = await file.read()
        contacts = []

        if filename.endswith('.csv'):
            decoded = contents.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded))
            for row in csv_reader:
                phone = row.get("phone_number") or row.get("phone") or row.get("Phone Number") or row.get("phonenumber")
                name = row.get("customer_name") or row.get("name") or row.get("Customer Name") or row.get("customername") or "Customer"
                r_ent = row.get("enterprise_id") or row.get("enterprise") or enterprise_id
                r_ag = row.get("agent_id") or row.get("agent") or agent_id
                r_cmp = row.get("campaign_id") or row.get("campaign") or batch_campaign_id
                
                if phone:
                    contacts.append({
                        "phone": str(phone).strip(),
                        "name": str(name).strip(),
                        "enterprise_id": r_ent,
                        "agent_id": r_ag,
                        "campaign_id": r_cmp
                    })
        else:
            # Excel parser using openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents), read_only=True)
            sheet = wb.active
            
            # Read header row
            first_row = next(sheet.iter_rows(max_row=1))
            headers = [cell.value for cell in first_row]
            
            phone_idx = None
            name_idx = None
            ent_idx = None
            ag_idx = None
            cmp_idx = None
            
            for idx, header in enumerate(headers):
                if header:
                    h = str(header).lower().strip()
                    if h in ["phone_number", "phone", "phone number", "phonenumber"]:
                        phone_idx = idx
                    elif h in ["customer_name", "name", "customer name", "customername"]:
                        name_idx = idx
                    elif h in ["enterprise_id", "enterprise", "enterprise id"]:
                        ent_idx = idx
                    elif h in ["agent_id", "agent", "agent id"]:
                        ag_idx = idx
                    elif h in ["campaign_id", "campaign", "campaign id"]:
                        cmp_idx = idx
                        
            # Fallbacks if headers are missing
            if phone_idx is None:
                phone_idx = 0
            if name_idx is None:
                name_idx = 1 if len(headers) > 1 else 0
                
            # Iterate data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= phone_idx:
                    continue
                phone = row[phone_idx]
                name = row[name_idx] if len(row) > name_idx else "Customer"
                r_ent = row[ent_idx] if ent_idx is not None and len(row) > ent_idx else enterprise_id
                r_ag = row[ag_idx] if ag_idx is not None and len(row) > ag_idx else agent_id
                r_cmp = row[cmp_idx] if cmp_idx is not None and len(row) > cmp_idx else batch_campaign_id
                
                if phone:
                    contacts.append({
                        "phone": str(phone).strip(),
                        "name": str(name).strip() if name else "Customer",
                        "enterprise_id": str(r_ent).strip() if r_ent else enterprise_id,
                        "agent_id": str(r_ag).strip() if r_ag else agent_id,
                        "campaign_id": str(r_cmp).strip() if r_cmp else batch_campaign_id
                    })

        initiated_calls = []
        for contact in contacts:
            phone = contact["phone"]
            name = contact["name"]
            
            # Call initiate_outbound_call controller function
            result = await call_controller.initiate_outbound_call(
                phone_number=phone,
                customer_name=name,
                enterprise_id=contact.get("enterprise_id"),
                agent_id=contact.get("agent_id"),
                campaign_id=contact.get("campaign_id")
            )
            initiated_calls.append({
                "phone_number": phone,
                "customer_name": name,
                "enterprise_id": result.get("enterprise_id"),
                "agent_id": result.get("agent_id"),
                "campaign_id": result.get("campaign_id"),
                "success": result.get("success", False),
                "call_sid": result.get("call_sid"),
                "call_sids": result.get("call_sids"),
                "error": result.get("error")
            })
            
        return {
            "success": True,
            "campaign_id": batch_campaign_id,
            "message": f"Successfully processed file under campaign '{batch_campaign_id}'. Initiated {len(initiated_calls)} calls.",
            "details": initiated_calls
        }
    except Exception as e:
        logger.error(f"Error triggering bulk calls: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Internal Server Error: {str(e)}"
        )

@router.get(
    "/campaigns",
    response_model=CampaignsResponse,
    summary="Get Enterprise & Agent Campaign Data",
    description="Retrieve all campaign data and call aggregations executed by an enterprise under a specific agent ID."
)
@router.get(
    "/outbound/campaigns",
    response_model=CampaignsResponse,
    include_in_schema=False
)
async def get_campaign_data(
    enterprise_id: str = Query(..., description="Enterprise ID to filter campaign data"),
    agent_id: Optional[str] = Query(None, description="Optional Agent ID to filter campaigns under the enterprise"),
    campaign_id: Optional[str] = Query(None, description="Optional specific Campaign ID to retrieve")
):
    """Retrieve campaign summaries and call details filtered by enterprise_id and optional agent_id."""
    result = await call_controller.fetch_campaign_data(
        enterprise_id=enterprise_id,
        agent_id=agent_id,
        campaign_id=campaign_id
    )
    
    if not result.get("success"):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=result.get("error", "An error occurred retrieving campaign data")
        )
        
    return result
